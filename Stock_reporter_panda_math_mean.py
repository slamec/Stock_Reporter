#!/usr/bin/python3
import pandas as pd 
import openpyxl

#read cdv from stock_reporter module
portfolio_read = pd.read_csv("portfolio_new.csv")

#check for duplicates and make meand, avoin indexes
portfolio = portfolio_read.groupby(['Symbol'], as_index= False).mean()
#sum of tickers quantity 
portfolio_sum = portfolio_read.groupby(['Symbol'], as_index= False).sum()

#used for quantity as number
total_quantity = portfolio_sum['Quantity']
#make gain/lose in %
divide_columns = (total_quantity * portfolio['Current price']) / (total_quantity * portfolio['Purchase price']) - 1
#count differences
difference = portfolio['Current price'] - portfolio['Purchase price']
#total absolute gain 
total_gain = (total_quantity * portfolio['Current price']) - (total_quantity * portfolio['Purchase price'])

current_price_sum = (total_quantity * portfolio['Current price'])
purchase_price_sum = (total_quantity * portfolio['Purchase price'])

#create new list with %
percent_list = []
for i in divide_columns:
    number = i*100
    format_number = "{:.2f} {}".format(number, "%")
    percent_list.append(format_number)

#create a list with differences
difference_list = []
for i in difference:
    diff_format = "{:.2f}".format(i)
    difference_list.append(diff_format)

#cerate new quantity list
quantity_list = []
for i in total_quantity:
    quantity_format = "{:.2f}".format(i)
    quantity_list.append(quantity_format)

#new list of total absolute gain
total_gain_list = []
for i in total_gain:
    gain_format = "{:.2f}".format(i)
    total_gain_list.append(gain_format)

current_price_sum_list = []
for i in current_price_sum:
    current_price_format = "{:.2f}".format(i)
    current_price_sum_list.append(current_price_format)

purchase_price_sum_list = []
for i in purchase_price_sum:
    purchase_price_format = "{:.2f}".format(i)
    purchase_price_sum_list.append(purchase_price_format)

#delete original 'Quantity' column
del portfolio['Quantity']

#convert list to floating numbers 
portfolio_sum_int_current = [int(float(i)) for i in current_price_sum_list]
portfolio_sum_int_purchase = [int(float(i)) for i in purchase_price_sum_list]

#list of positive and negative numbers from the list
portfolio_sum_positive = [i for i in portfolio_sum_int_current if i >= 0]
portfolio_sum_negative = [i for i in portfolio_sum_int_current if i < 0]
portfolio_sum_positive_purchase = [i for i in portfolio_sum_int_purchase if i >= 0]
portfolio_sum_negative_purchase = [i for i in portfolio_sum_int_purchase if i < 0]


portfolio_total_gain_current =  sum(portfolio_sum_negative) + sum(portfolio_sum_positive)
portfolio_total_gain_purchase = sum(portfolio_sum_negative_purchase) + sum(portfolio_sum_positive_purchase)

print(portfolio_total_gain_current)
print(portfolio_total_gain_purchase)

total_gain_percent = (1 - (portfolio_total_gain_purchase / portfolio_total_gain_current)) * 100  

print(float(total_gain_percent))

#make new rows to a csv/excel
portfolio["Quantity"] = quantity_list
portfolio["Current Difference"] = difference_list
portfolio["Total gain"] = total_gain_list
portfolio["Total gain (%)"] = percent_list
portfolio["P/L (%)"] = "{} {:.2f}%".format("P/L ", float(total_gain_percent))

#write everything to Excel
portfolio.to_excel("Stock_final.xlsx", index = None, header=True)
portfolio.to_csv("Stock_final.csv", index = False)


def count_rows():
    """Counts rows in a given dataframe"""
    df = pd.DataFrame(portfolio)
    result = df.count(numeric_only=True)
    #returns number of rows plus 3
    return result[1] + 3

def merge_cells(file_name):
    """Merge cells in Excel file based given coordinates"""
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    coordinate_a = "H2:"
    coordinate_b = "H" + str(count_rows()) 
    coordinates = str(coordinate_a) + str(coordinate_b)

    sheet.merge_cells(coordinates)
    wb.save("Stock_final.xlsx")

#call the function merging cells
merge_cells("Stock_final.xlsx")

df = pd.read_excel("Stock_final.xlsx")
df.to_html("Stock_final.html")
